#!/usr/bin/env python3
"""
Helper script to update both CSV and Excel files with formulas
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

def update_metadata_files():
    """Read CSV and update Excel file, preserving formulas"""

    # Read CSV
    df = pd.read_csv('corpus_metadata.csv')

    # Load existing workbook to preserve formulas
    try:
        wb = load_workbook('corpus_metadata.xlsx')
        ws = wb.active
    except FileNotFoundError:
        print("Excel file not found, will create new one")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Corpus"

    # Clear data rows (but keep formulas below)
    ws.delete_rows(1, 55)  # Delete first 55 rows (header + 53 poems + buffer)

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, 1):
            # Handle TRUE/FALSE as boolean
            if col_idx == 35:  # collected column (position 35 after adding new metadata)
                val = True if str(value).upper() == 'TRUE' else (False if str(value).upper() == 'FALSE' else value)
                ws.cell(row=row_idx+2, column=col_idx, value=val)
            else:
                ws.cell(row=row_idx+2, column=col_idx, value=value)

    # Re-add formulas if they don't exist (starting at row 57)
    start_row = 57

    # Check if formulas exist
    if ws[f'A{start_row}'].value != 'SUMMARY STATISTICS (COLLECTED POEMS ONLY)':
        print("Adding formulas...")
        ws[f'A{start_row}'] = 'SUMMARY STATISTICS (COLLECTED POEMS ONLY)'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)

        stats_row = start_row + 2

        # Period
        ws[f'A{stats_row}'] = 'PERIOD'
        ws[f'B{stats_row}'] = 'Count'
        periods = ['Tudor', 'Elizabethan', 'Jacobean', 'Caroline', 'Interregnum',
                   'Restoration', 'Neoclassical', 'Romantic', 'Victorian', 'Modernist',
                   'Postwar', 'Contemporary']
        for i, period in enumerate(periods):
            row = stats_row + 1 + i
            ws[f'A{row}'] = period
            ws[f'B{row}'] = f'=COUNTIFS(E:E,A{row},AI:AI,TRUE)'

        # Mode
        ws[f'D{stats_row}'] = 'LITERARY_MOVEMENT'
        ws[f'E{stats_row}'] = 'Count'
        movements = ['Renaissance', 'Metaphysical', 'Augustan', 'Graveyard School',
                     'Romanticism', 'Pre-Raphaelite', 'Imagism', 'Modernism',
                     'Harlem Renaissance', 'Beat', 'Confessional', 'Black Arts',
                     'Language Poetry']
        for i, movement in enumerate(movements):
            row = stats_row + 1 + i
            ws[f'D{row}'] = movement
            ws[f'E{row}'] = f'=COUNTIFS(F:F,D{row},AI:AI,TRUE)'

        # Mode
        ws[f'G{stats_row}'] = 'MODE'
        ws[f'H{stats_row}'] = 'Count'
        modes = ['Lyric', 'Narrative', 'Dramatic', 'Mixed']
        for i, mode in enumerate(modes):
            row = stats_row + 1 + i
            ws[f'G{row}'] = mode
            ws[f'H{row}'] = f'=COUNTIFS(H:H,G{row},AI:AI,TRUE)'

        # Stance
        ws[f'J{stats_row}'] = 'STANCE'
        ws[f'K{stats_row}'] = 'Count'
        stances = ['Apostrophic', 'Meditative', 'Descriptive', 'Argumentative',
                   'Narrative', 'Satiric', 'Prophetic', 'Ceremonial']
        for i, stance in enumerate(stances):
            row = stats_row + 1 + i
            ws[f'J{row}'] = stance
            ws[f'K{row}'] = f'=COUNTIFS(K:K,"*{stance}*",AI:AI,TRUE)'

        # Rhetorical mode
        ws[f'M{stats_row}'] = 'RHETORICAL_MODE'
        ws[f'N{stats_row}'] = 'Count'
        for i, rhet in enumerate(['Epideictic', 'Deliberative', 'Forensic', '(blank)']):
            row = stats_row + 1 + i
            ws[f'M{row}'] = rhet
            if rhet == '(blank)':
                ws[f'N{row}'] = f'=COUNTIFS(L:L,"",AI:AI,TRUE)'
            else:
                ws[f'N{row}'] = f'=COUNTIFS(L:L,M{row},AI:AI,TRUE)'

        # Length statistics
        length_row = stats_row + 10
        ws[f'P{length_row}'] = 'LENGTH STATISTICS'
        ws[f'P{length_row + 1}'] = 'Total collected'
        ws[f'Q{length_row + 1}'] = '=COUNTIF(AI:AI,TRUE)'
        ws[f'P{length_row + 2}'] = 'Mean lines'
        ws[f'Q{length_row + 2}'] = '=AVERAGEIF(AI:AI,TRUE,AG:AG)'
        ws[f'P{length_row + 3}'] = 'Median lines'
        ws[f'Q{length_row + 3}'] = '=MEDIAN(IF(AI2:AI54=TRUE,AG2:AG54))'
        ws[f'P{length_row + 4}'] = 'Min lines'
        ws[f'Q{length_row + 4}'] = '=MINIFS(AG:AG,AI:AI,TRUE)'
        ws[f'P{length_row + 5}'] = 'Max lines'
        ws[f'Q{length_row + 5}'] = '=MAXIFS(AG:AG,AI:AI,TRUE)'
        ws[f'P{length_row + 6}'] = 'Mean words'
        ws[f'Q{length_row + 6}'] = '=AVERAGEIF(AI:AI,TRUE,AH:AH)'
        ws[f'P{length_row + 7}'] = 'Median words'
        ws[f'Q{length_row + 7}'] = '=MEDIAN(IF(AI2:AI54=TRUE,AH2:AH54))'

    # Save
    wb.save('corpus_metadata.xlsx')
    print("✓ Updated corpus_metadata.xlsx")

if __name__ == '__main__':
    update_metadata_files()
